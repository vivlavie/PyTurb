#Read Turbulent over the helideck cylinder of 30m
#To read heat on and around SCE's

#Sigma at each point -> average of sigma -> max sigma in the air space
import math
import numpy as np
import os
# import dill
from kfxtools import * #fit for Python 2.

# Vmax = 17.5
# basefolder = "./"

Vmax = 18.8
basefolder = "./18.8/"


sigma_cri = 1.75
c_iso = 1.5 # 6.29% unaavailability
# c_iso = 2.22 #1.96% unavilabilty
k_cri = c_iso * sigma_cri**2
# k_cri = 6.80



fieldnum = 0

Vd = ['Stern','-150','-120','PORT','-60','-30','Bow','30','60','STBD','120','150' ]
Ts = ['T01','T12', 'T11','T10','T09','T08','T07','T06','T05','T04','T03','T02']

from openpyxl import load_workbook
iWR=load_workbook(filename='WindRose_v02.xlsx',data_only=True)
shWR = iWR['Sheet1']

Pwind = np.zeros((24,8))
VwRange = np.arange(0,22.5,2.5)
for row in range(37,61,1):
    for col in range(13,21,1):
        Pwind[row-37,col-13] = float(shWR.cell(row,col).value)

Pwind2 = np.zeros((12,8))
Pwind2[0,:] = Pwind[2*0,:] + 0.5*(Pwind[2*12-1,:]+Pwind[2*0+1,:])
for row in range(1,12):
    Pwind2[row,:] = Pwind[2*row,:] + 0.5*(Pwind[2*row-1,:]+Pwind[2*row+1,:])

Pwind = Pwind2


Xs = np.arange(5,27,2)
Ys = np.arange(22,44,2)
Zs = np.arange(64.5,94.5,5)
Ks = np.zeros(len(Zs))
s = np.zeros(len(Zs))
MH = np.zeros(len(Ts))
TKE = np.zeros(len(Ts))
sigma = np.zeros(len(Ts))
V = np.zeros(len(Ts))
ti = 0


U = 0.
Uwdz = np.zeros((len(Ts),len(Zs)))
print "{:10s}".format("Height"), " ".join(map("{:8.2f}".format,Zs))
for t in Ts:
# for t in ['T05']:    
    fdr = t    
    fn = basefolder + "/" + fdr+ "/" + t + "_tke_exit.r3d"        
    
    if (os.path.exists(fn) == False):
        print(fn + " does not exist")    
    else:    
        #Rad radiation
        T = readr3d(fn)            
        fieldname = T.names[fieldnum]
        # print(t,fn,fieldname)

        zi = 0
        kmax = 0
        for z in Zs:
            ksum = 0.
            s_sum = 0.
            for y in Ys:
                for x in Xs:
                    k = T.point_value(x,y,z,fieldnum)
                    ksum += k
                    s_sum += np.sqrt(k/c_iso)
                    if k > kmax:
                        MH[ti] = z
                        kmax = k
                        s_max = np.sqrt(k/c_iso)

            # Ks[zi] = ksum/(len(Xs)*len(Ys)) #average of k
            # print("Z= {:6.1f} TKE_avg = {:6.2f}".format(z, Ks[zi]))
            # Vzi = min(Vmax,Vmax*sigma_cri/np.sqrt(Ks[zi]/c_iso))

            s[zi] = s_sum/(len(Xs)*len(Ys)) #average of sigma
            # print("Z= {:6.1f} sigma_avg = {:6.2f}".format(z, s[zi]))
            Vzi = min(Vmax,Vmax*sigma_cri/s[zi])
            # print(Vzi,Vzi2)
            Uti = 0.
            for vi in range(0,len(VwRange)-1):
                if (Vzi > VwRange[vi]) and (Vzi <= VwRange[vi+1]):
                    Uti += (VwRange[vi+1]-Vzi)/2.5*Pwind[ti,vi]
                elif (Vzi < VwRange[vi]):
                    Uti += Pwind[ti,vi]            
            Uwdz[ti,zi] = Uti #Unavailability per wind direction and elevation
            zi += 1
        # print "{:10s}".format(Vd[Ts.index(t)]), " ".join(map("{:8.2f}".format,Ks))
        # print "{:10s}".format(Vd[Ts.index(t)]), " ".join(map("{:8.2f}".format,np.sqrt(Ks/c_iso)))
        print "{:10s}".format(Vd[Ts.index(t)]), " ".join(map("{:8.2f}".format,s))
        # TKE[ti] = max(Ks)
        sigma[ti] = max(s)
        # V[ti] = min(Vmax,Vmax*k_cri/TKE[ti]) #This is wrong in that 'linear scaling' is applined to k [m2/s2]. #unavailability 24%
        # V[ti] = min(Vmax,Vmax*sigma_cri/np.sqrt(TKE[ti]/c_iso))
        V[ti] = min(Vmax,Vmax*sigma_cri/sigma[ti])

        Uti = 0.
        for vi in range(0,len(VwRange)-1):
            if (V[ti] > VwRange[vi]) and (V[ti] <= VwRange[vi+1]):
                Uti += (VwRange[vi+1]-V[ti])/2.5*Pwind[ti,vi]
            elif (V[ti] < VwRange[vi]):
                Uti += Pwind[ti,vi]
        U += Uti
        print("{:10s} {:6.2f} {:6.2f}% {:6.2f}%".format(Vd[ti],V[ti],100*Uti, 100*U))
        # print("TKE_Max = {:6.2f}".format(max(Ks)))
        # print("Scenario: {:s} {:6s} Max speed for Turbulence: {:6.1f} Max turbulence height {:6.1f} Max turbulence {:6.1f} Max(Ks): {:6.1f}".format(t, Vd[ti],min(Vmax,Vmax*k_cri/max(Ks)),MH[ti],kmax,max(Ks)))
    ti += 1

#
""" 
ti = 0
print "{:10s}".format("Height"), " ".join(map("{:9.2f}".format,Zs))
for t in Ts:
    print "{:10s}".format(Vd[ti])," ".join(map("{:8.2f}%".format,100*Uwdz[ti,:]))
    ti += 1


for d in Vd:
    print "{:10s}".format(d), " ".join(map("{:9.2f}%".format,100*Pwind2[Vd.index(d),:])) """